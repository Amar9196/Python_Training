import xlwt
from xlwt import Workbook, Worksheet
import psycopg2
from openpyxl.reader.excel import load_workbook

workbook = load_workbook(r"connectionxl.xlsx")
connection_sheet = workbook["connection"]

dbname = connection_sheet.cell(row=2, column=1).value
user = connection_sheet.cell(row=2, column=2).value
host = connection_sheet.cell(row=2, column=3).value
password = connection_sheet.cell(row=2, column=4).value

wb = xlwt.Workbook('student_details.xlsx')
sheet = wb.add_sheet('Sheet1')
try:
    conn = psycopg2.connect(f"dbname='{dbname}' user='{user}' host='{host}' password='{password}'")
    print(conn)

except:
    print("Unable to connect to database")

try:
    cur = conn.cursor()
    cur.execute("""select sid,name,age from student""")

    rows = cur.fetchall()

    print("Retrieved data")
    i = 0
    for row in rows:
        sheet.write(i, 0, row[0])
        sheet.write(i, 1, row[1])
        sheet.write(i, 2, row[2])
        i = i + 1

except Exception as ex:
    print("Error in fetching data", ex)

finally:
    conn.close()
    wb.save('student_details.xlsx')